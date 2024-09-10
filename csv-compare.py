import openpyxl as op

def compare_excel_sheets(file1, file2, sheet_name):
    # Load the two workbooks
    wb1 = op.load_workbook(file1)
    wb2 = op.load_workbook(file2)
    
    # Load the specific sheet by name from both workbooks
    sheet1 = wb1[sheet_name]
    sheet2 = wb2[sheet_name]

    # Get the dimensions of the sheets
    max_row = max(sheet1.max_row, sheet2.max_row)
    max_column = max(sheet1.max_column, sheet2.max_column)

    # Iterate over each row and column
    for row in range(1, max_row + 1):
        for col in range(1, max_column + 1):
            cell1 = sheet1.cell(row=row, column=col).value
            cell2 = sheet2.cell(row=row, column=col).value

            # Compare the cells and print if different
            if cell1 != cell2:
                print(f"Difference found at Row {row}, Column {col}:")
                print(f"{file1} value: {cell1}")
                print(f"{file2} value: {cell2}\n")

filePath1 = input("Path to first file: ")
filePath2 = input("Path to second file: ")
sheetname = input("Sheetname for compare: ")

compare_excel_sheets(filePath1, filePath2, sheetname)
